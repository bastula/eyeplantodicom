#!/usr/bin/env python
# -*- coding: utf-8 -*-
# eyeplantodicom.py
"""Convert an Eyeplan Excel dose file to a DICOM RT Dose file."""
# Copyright (c) 2015 Aditya Panchal

import dicom
import openpyxl
import pandas as pd
import numpy as np
from scipy import interpolate
import logging


class EyeplanToDICOM(object):
    """Class that reads and converts an Eyeplan Excel dose file
       to a DICOM RT Dose file"""
    def __init__(self, filename, sourcedicom):

        self.logger = logging.getLogger('eyeplantodicom')

        self.filename = filename
        self.sourcedicom = sourcedicom

        self.read_excel_file()
        self.read_dicom_file()

    def read_excel_file(self):
        """Read the Eyeplan Excel file from disk."""

        # Get the patient info
        wb = openpyxl.load_workbook(filename=self.filename)
        ws = wb.get_active_sheet()
        self.patientname = str(ws['A1'].value)
        self.patientid = str(ws['A2'].value)

        # Skip the first two rows as it contains patient specific info
        self.df = pd.io.excel.read_excel(self.filename, skiprows=[0, 1])

        return self.df

    def read_dicom_file(self):
        """Read the example DICOM RT Dose file from disk."""

        self.rd = dicom.read_file(self.sourcedicom, force=True)

    def convert_data(self, progressfunc=None, updatefunc=None, dryrun=False):
        """Convert the Eyeplan Excel data to DICOM RT Dose"""

        # Get the y planes to sort on
        yplanes = self.df.Y.unique()
        self.logger.debug("Y planes: %g %s", yplanes.shape[0], yplanes)

        # Get some basic grid size information
        self.logger.debug(
            "xmin: %g, xmax: %g, zmin: %g, zmax: %g",
            self.df.X.min(),
            self.df.X.max(),
            self.df.Z.min(),
            self.df.Z.max())

        self.pixel_spacing = [0.1, 0.1]

        planes = []
        # Iterate through all the y planes and interpolate each one
        for y in np.nditer(yplanes):
            self.logger.debug("Processing Y plane slice # %g mm", y)
            yplane = self.df[self.df.Y == [y]][['X', 'Z', 'Dose']]

            # Pivot the data frame such that the dose values are in a grid
            pivot = yplane.pivot(index='Z', columns='X', values='Dose')

            x = pivot.columns
            z = pivot.index
            # Create a RectBivariateSpline to interpolate the data
            rbs = interpolate.RectBivariateSpline(z, x, pivot)
            interpolation = rbs(
                np.arange(z[0], z[-1], self.pixel_spacing[1]),
                np.arange(x[0], x[-1], self.pixel_spacing[0]),
                grid=True)
            planes.append(interpolation.T)

        # Flip the dose grid so that the x axis is L-R, z axis is U-D
        dosegrid = np.dstack(planes).T
        # Scale the dose grid to fit within 2^32-1 (32-bit unsigned int)
        scale = 8
        dosegrid = dosegrid * pow(10, scale)
        # Remove all values less than desired threshold
        dosegrid[dosegrid < 0.001] = 0

        # Update the source DICOM file with the new dose information
        self.rd.ImagePositionPatient = \
            [str(self.df.X.min()),
             str(self.df.Z.min()),
             str(self.df.Y.min())]
        self.rd.GridFrameOffsetVector = np.sort(
            (yplanes - yplanes[-1])).tolist()
        self.rd.DoseGridScaling = pow(10, -scale)
        self.rd.PixelSpacing = self.pixel_spacing
        # Convert the dose grid to unsigned integer 32-bit
        self.rd.PixelData = np.uint32(dosegrid).tostring()
        self.rd.Rows = dosegrid.shape[1]
        self.rd.Columns = dosegrid.shape[2]
        self.rd.NumberOfFrames = dosegrid.shape[0]
        if 'DVHs' in self.rd:
            del self.rd.DVHs

        # Update the UIDs
        self.rd.SOPInstanceUID = dicom.UID.generate_uid(None)
        self.rd.StudyInstanceUID = dicom.UID.generate_uid(None)
        self.rd.SeriesInstanceUID = dicom.UID.generate_uid(None)
        self.rd.FrameOfReferenceUID = dicom.UID.generate_uid(None)

        # Update the patient name
        self.rd.PatientName = self.patientname
        self.rd.PatientID = self.patientid

        return self.rd

if __name__ == '__main__':

    import sys
    import argparse
    logger = logging.getLogger('eyeplantodicom')
    logger.setLevel(logging.INFO)
    ch = logging.StreamHandler()
    ch.setFormatter(logging.Formatter('%(levelname)s: %(message)s'))
    logger.addHandler(ch)

    # Set up argparser to parse the command-line arguments
    class DefaultParser(argparse.ArgumentParser):
        def error(self, message):
            sys.stderr.write('error: %s\n' % message)
            self.print_help()
            sys.exit(2)

    parser = DefaultParser(
        description="Convert an Eyeplan Excel dose file to a DICOM RT " +
                    "Dose file.")
    parser.add_argument("filename",
                        help="Excel (.xlsx) file name")
    parser.add_argument("sourcedicom",
                        help="Source DICOM RT Dose file name")
    parser.add_argument("outputdicom",
                        help="Output DICOM RT Dose file name")
    parser.add_argument("-d", "--debug",
                        help="Show debug log",
                        action="store_true")

    # If there are no arguments, display help and exit
    if len(sys.argv) == 1:
        parser.print_help()
        sys.exit(0)
    args = parser.parse_args()

    # Set debug logging if the debug flag is set
    if args.debug:
        logger.setLevel(logging.DEBUG)

    # Read the Excel file
    eyeplan = EyeplanToDICOM(args.filename, args.sourcedicom)
    # Convert the data
    rd = eyeplan.convert_data()
    # Save the output DICOM RT file to disk
    rd.save_as(args.outputdicom)
